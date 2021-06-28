import xml.etree.ElementTree as ET
from openpyxl.styles import Border, Side, Alignment, Font
import openpyxl
import shutil
from num2words import num2words


def getContent(file):
    def getChilds(root):
        d = {}
        index = 0
        for child in root:
            flag = False
            if child.tag and child.attrib:
                if child.tag not in d:
                    d[child.tag] = child.attrib
                else:
                    if not flag:
                        d[child.tag] = [d[child.tag]]
                        flag = True
                        index = 1
                        d[child.tag].append(child.attrib)
                    else:
                        d[child.tag].append(child.attrib)
                        index += 1

            try:
                temp = getChilds(child)
                for key in temp.keys():
                    if not flag:
                        d[child.tag][key] = temp[key]
                    else:
                        d[child.tag][index][key] = temp[key]
            except IndexError:
                pass
        return d

    tree = ET.parse(file)
    root = tree.getroot()[0]
    content = {root.tag: root.attrib}
    for child in root:
        if getChilds(child):
            if child.tag not in content:
                content[child.tag] = []
            content[child.tag].append(getChilds(child))
        else:
            if child.tag not in content:
                content[child.tag] = []
            content[child.tag].append(child.attrib)

    return content


def putIntoXLS(data):
    k = 13
    thin = 'thin'
    thick = 'medium'
    shutil.copyfile('ЭДО\\Schet.xlsx', 'ЭДО\\result.xlsx')
    workbook = openpyxl.load_workbook('ЭДО\\Schet.xlsx')
    sheet = workbook['TDSheet']

    sheet['B2'] = data['Поставщик'][0]['БанкРекв']['НаимБанк']
    sheet['AD2'] = data['Поставщик'][0]['БанкРекв']['БИК']
    sheet['AD3'] = data['Поставщик'][0]['БанкРекв']['РСчет']
    sheet['AD5'] = data['Поставщик'][0]['БанкРекв']['КСчет']
    sheet['B6'] = data['Поставщик'][0]['СвЮЛ']['Название']
    sheet['B10'] = f'{data["Документ"]["Название"]} №{data["Документ"]["Номер"]} от {data["Документ"]["Дата"]}  г.'
    sheet['G14'] = data['Поставщик'][0]['СвЮЛ']['Название'] + ',' + data['Поставщик'][0]['Адрес']['АдрТекст']
    sheet['G17'] = data['Покупатель'][0]['СвЮЛ']['Название'] + ',' + data['Покупатель'][0]['Адрес']['АдрТекст']
    sheet['G20'] = data['Параметр'][0]['Значение']
    row = 23
    i = 1
    arr = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W',
           'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO',
           'AP', 'AQ']

    for el in data['ТаблДок'][0]['СтрТабл']:
        sheet.merge_cells(f'B{row}:C{row}')
        sheet.merge_cells(f'D{row}:X{row}')
        sheet.merge_cells(f'Y{row}:AB{row}')
        sheet.merge_cells(f'AC{row}:AE{row}')
        sheet.merge_cells(f'AF{row}:AJ{row}')
        sheet.merge_cells(f'AK{row}:AQ{row}')
        sheet.row_dimensions[row].height = 50
        sheet[f'B{row}'] = i
        sheet[f'D{row}'] = el['Название']
        try:
            sheet[f'AF{row}'] = el['Цена']
        except KeyError:
            sheet[f'AF{row}'] = el['Сумма']
        sheet[f'AK{row}'] = el['Сумма']

        for letter in arr:
            if letter == 'B':
                sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                        right=Side(border_style=thin, color='FF000000'),
                                                        bottom=Side(border_style=thin, color='FF000000'),
                                                        left=Side(border_style=thick, color='FF000000'))
            elif letter == 'AQ':
                sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                        right=Side(border_style=thin, color='FF000000'),
                                                        bottom=Side(border_style=thin, color='FF000000'),
                                                        left=Side(border_style=thick, color='FF000000'))
            else:
                sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                        right=Side(border_style=thin, color='FF000000'),
                                                        bottom=Side(border_style=thin, color='FF000000'),
                                                        left=Side(border_style=thin, color='FF000000'))

            sheet[f'{letter}{row}'].alignment = Alignment(horizontal='left',
                                                          vertical='top',
                                                          text_rotation=0,
                                                          wrap_text=True,
                                                          shrink_to_fit=False, indent=0)

        i += 1
        row += 1

    row -= 1
    for letter in arr:
        if letter == 'B':
            sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                    right=Side(border_style=thin, color='FF000000'),
                                                    bottom=Side(border_style=thick, color='FF000000'),
                                                    left=Side(border_style=thick, color='FF000000'))
        elif letter == 'AQ':
            sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                    right=Side(border_style=thin, color='FF000000'),
                                                    bottom=Side(border_style=thick, color='FF000000'),
                                                    left=Side(border_style=thick, color='FF000000'))
        else:
            sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                    right=Side(border_style=thin, color='FF000000'),
                                                    bottom=Side(border_style=thick, color='FF000000'),
                                                    left=Side(border_style=thin, color='FF000000'))

    row += 2
    for i in range(3):
        sheet.merge_cells(f'AL{row + i}:AQ{row + i}')
        sheet[f'AK{row + i}'].alignment = Alignment(horizontal='right',
                                                    vertical='top',
                                                    text_rotation=0,
                                                    wrap_text=False,
                                                    shrink_to_fit=False, indent=0)
        sheet[f'AL{row + i}'].alignment = Alignment(horizontal='right',
                                                    vertical='top',
                                                    text_rotation=0,
                                                    wrap_text=False,
                                                    shrink_to_fit=False, indent=0)
        sheet.row_dimensions[row + i].height = 15
    sheet[f'AK{row}'] = 'Итого:'
    sheet[f'AL{row}'] = data['ТаблДок'][0]['ИтогТабл']['Сумма']
    row += 1
    sheet[f'AK{row}'] = 'В том числе НДС:'
    sheet[f'AL{row}'] = data['ТаблДок'][0]['ИтогТабл']['НДС']['Сумма']
    row += 1
    sheet[f'AK{row}'] = 'Всего к оплате:'
    sheet[f'AL{row}'] = data['ТаблДок'][0]['ИтогТабл']['Сумма']

    row += 1
    sheet.row_dimensions[row].height = k
    rubles = int(sheet[f"AL{row - 1}"].value.split('.')[0])
    cents = sheet[f"AL{row - 1}"].value.split('.')[1]
    rubles = num2words(rubles, to='cardinal', lang='ru')
    sheet.merge_cells(f'B{row}:AR{row}')
    sheet[f'B{row}'] = f'Всего наименований {i} на сумму {sheet[f"AL{row - 1}"].value} руб'
    sheet[f'B{row}'].font = Font(name='Arial',
                                 size=9,
                                 bold=False,
                                 italic=False,
                                 vertAlign=None,
                                 underline='none',
                                 strike=False,
                                 color='FF000000')

    row += 1
    sheet.row_dimensions[row].height = k
    sheet.merge_cells(f'B{row}:AP{row}')
    sheet[f'B{row}'] = f'{rubles.capitalize()} рублей {cents} копеек'
    sheet[f'B{row}'].font = Font(name='Arial',
                                 size=9,
                                 bold=True,
                                 italic=False,
                                 vertAlign=None,
                                 underline='none',
                                 strike=False,
                                 color='FF000000')

    row += 2
    index = 0
    text = ['Внимание!', 'Оплата данного счета означает согласие с условиями поставки товара.',
            'Уведомление об оплате обязательно, в противном случае не гарантируется наличие товара на складе.',
            'Товар отпускается по факту прихода денег на р/с Поставщика, самовывозом, при наличии доверенности и паспорта.']

    for i in range(4):
        sheet.row_dimensions[row + i].height = k
        sheet.merge_cells(f'B{row + i}:AQ{row + i}')
        sheet[f'B{row + i}'] = text[index]
        sheet[f'B{row + i}'].font = Font(name='Arial',
                                         size=9,
                                         bold=False,
                                         italic=False,
                                         vertAlign=None,
                                         underline='none',
                                         strike=False,
                                         color='FF000000')
        sheet[f'B{row + i}'].alignment = Alignment(horizontal='left',
                                                   vertical='top',
                                                   text_rotation=0,
                                                   wrap_text=True,
                                                   shrink_to_fit=False, indent=0)

        index += 1
    row += 3
    sheet.row_dimensions[row].height = 2 * k
    row += 1

    for letter in arr + ['AQ']:
        sheet[f'{letter}{row}'].border = Border(bottom=Side(border_style=thick, color='FF000000'))

    row += 2

    sheet[f'B{row}'] = 'Руководитель'
    sheet[f'B{row}'].font = Font(name='Arial',
                                 size=9,
                                 bold=True,
                                 italic=False,
                                 vertAlign=None,
                                 underline='none',
                                 strike=False,
                                 color='FF000000')
    sheet[f'B{row}'].alignment = Alignment(horizontal='left',
                                           vertical='bottom',
                                           text_rotation=0,
                                           wrap_text=False,
                                           shrink_to_fit=False, indent=0)
    sheet[f'Z{row}'] = 'Бухгалтер'
    sheet[f'Z{row}'].font = Font(name='Arial',
                                 size=9,
                                 bold=True,
                                 italic=False,
                                 vertAlign=None,
                                 underline='none',
                                 strike=False,
                                 color='FF000000')
    sheet[f'Z{row}'].alignment = Alignment(horizontal='left',
                                           vertical='bottom',
                                           text_rotation=0,
                                           wrap_text=False,
                                           shrink_to_fit=False, indent=0)

    letter_underline = ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'AG', 'AH', 'AI',
                        'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS']

    for i in letter_underline:
        sheet[f'{i}{row}'].border = Border(bottom=Side(border_style=thin, color='FF000000'))

    workbook.save('ЭДО\\result.xlsx')
