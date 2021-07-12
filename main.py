import sys
from PyQt5 import uic, QtWidgets
from PyQt5.QtWidgets import QMainWindow, QFileDialog
import xml.etree.ElementTree as ET
from openpyxl.styles import Border, Side, Alignment, Font
import shutil
from openpyxl import load_workbook, Workbook
from num2words import num2words
from datetime import datetime


def convert(open_filename, result_file='result'):
    def create_table_end(start_row, cols, tree):
        arr = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
               'W',
               'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO',
               'AP', 'AQ']

        path = 'Документ,ТаблДок,СтрТабл'.split(',')
        root = tree.getroot()
        for step in path[:-1]:
            a = [x.tag for x in root]
            root = root[a.index(step)]

        count = [x.tag for x in root].count('СтрТабл')

        for i in range(1, count + 1):
            row = start_row + i
            sheet.merge_cells(f'B{row}:C{row}')
            sheet.merge_cells(f'D{row}:X{row}')
            sheet.merge_cells(f'Y{row}:AB{row}')
            sheet.merge_cells(f'AC{row}:AE{row}')
            sheet.merge_cells(f'AF{row}:AJ{row}')
            sheet.merge_cells(f'AK{row}:AQ{row}')
            sheet.row_dimensions[row].height = 50
            for col in cols:
                if col[0] == 'Номер':
                    sheet[f'{col[1]}{row}'] = i
                else:
                    sheet[f'{col[1]}{row}'] = get_root(col[0], tree, i - 1)

            for letter in arr:
                if letter == 'B':
                    sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                            right=Side(border_style=thin, color='FF000000'),
                                                            bottom=Side(border_style=thin, color='FF000000'),
                                                            left=Side(border_style=thick, color='FF000000'))
                elif letter == 'AQ':
                    sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                            right=Side(border_style=thin, color='FF000000'),
                                                            bottom=Side(border_style=thin, color='FF000000'),
                                                            left=Side(border_style=thick, color='FF000000'))
                else:
                    sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                            right=Side(border_style=thin, color='FF000000'),
                                                            bottom=Side(border_style=thin, color='FF000000'),
                                                            left=Side(border_style=thin, color='FF000000'))

                sheet[f'{letter}{row}'].alignment = Alignment(horizontal='left',
                                                              vertical='top',
                                                              text_rotation=0,
                                                              wrap_text=True,
                                                              shrink_to_fit=False, indent=0)

            i += 1
            for letter in arr:
                if letter == 'B':
                    sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                            right=Side(border_style=thin, color='FF000000'),
                                                            bottom=Side(border_style=thin, color='FF000000'),
                                                            left=Side(border_style=thick, color='FF000000'))
                elif letter == 'AQ':
                    sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                            right=Side(border_style=thick, color='FF000000'),
                                                            bottom=Side(border_style=thin, color='FF000000'),
                                                            left=Side(border_style=thick, color='FF000000'))
                else:
                    sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                            right=Side(border_style=thin, color='FF000000'),
                                                            bottom=Side(border_style=thin, color='FF000000'),
                                                            left=Side(border_style=thin, color='FF000000'))

        for letter in arr:
            if letter == 'B':
                sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                        right=Side(border_style=thin, color='FF000000'),
                                                        bottom=Side(border_style=thick, color='FF000000'),
                                                        left=Side(border_style=thick, color='FF000000'))
            elif letter == 'AQ':
                sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                        right=Side(border_style=thick, color='FF000000'),
                                                        bottom=Side(border_style=thick, color='FF000000'),
                                                        left=Side(border_style=thick, color='FF000000'))
            else:
                sheet[f'{letter}{row}'].border = Border(top=Side(border_style=thin, color='FF000000'),
                                                        right=Side(border_style=thin, color='FF000000'),
                                                        bottom=Side(border_style=thick, color='FF000000'),
                                                        left=Side(border_style=thin, color='FF000000'))
        row += 3
        for i in range(3):
            sheet.merge_cells(f'AL{row + i}:AQ{row + i}')
            sheet[f'AK{row + i}'].alignment = Alignment(horizontal='right',
                                                        vertical='top',
                                                        text_rotation=0,
                                                        wrap_text=False,
                                                        shrink_to_fit=False, indent=0)

            sheet[f'AL{row + i}'].alignment = Alignment(horizontal='right',
                                                        vertical='top',
                                                        text_rotation=0,
                                                        wrap_text=False,
                                                        shrink_to_fit=False, indent=0)
            sheet.row_dimensions[row + i].height = 15

        sheet[f'AK{row}'] = 'Итого:'
        sheet[f'AK{row}'].font = Font('Arial', 9, bold=True)
        sheet[f'AL{row}'].font = Font('Arial', 9, bold=True)
        sheet[f'AL{row}'] = get_root(f[-1], tree)
        row += 1
        sheet[f'AK{row}'].font = Font('Arial', 9, bold=True)
        sheet[f'AL{row}'].font = Font('Arial', 9, bold=True)
        sheet[f'AK{row}'] = 'В том числе НДС:'
        sheet[f'AL{row}'] = get_root(f[-2], tree)
        row += 1
        sheet[f'AK{row}'].font = Font('Arial', 9, bold=True)
        sheet[f'AL{row}'].font = Font('Arial', 9, bold=True)
        sheet[f'AK{row}'] = 'Всего к оплате:'
        sheet[f'AL{row}'] = get_root(f[-1], tree)

        row += 1
        sheet.row_dimensions[row].height = k
        rubles = int(sheet[f"AL{row - 1}"].value.split('.')[0])
        cents = sheet[f"AL{row - 1}"].value.split('.')[1]
        rubles = num2words(rubles, to='cardinal', lang='ru')
        sheet.merge_cells(f'B{row}:AR{row}')
        sheet[f'B{row}'] = f'Всего наименований {i} на сумму {sheet[f"AL{row - 1}"].value} руб'
        sheet[f'B{row}'].font = Font(name='Arial',
                                     size=9,
                                     bold=False,
                                     italic=False,
                                     vertAlign=None,
                                     underline='none',
                                     strike=False,
                                     color='FF000000')

        row += 1
        sheet.row_dimensions[row].height = k
        sheet.merge_cells(f'B{row}:AP{row}')
        sheet[f'B{row}'] = f'{rubles.capitalize()} рублей {cents} копеек'
        sheet[f'B{row}'].font = Font(name='Arial',
                                     size=9,
                                     bold=True,
                                     italic=False,
                                     vertAlign=None,
                                     underline='none',
                                     strike=False,
                                     color='FF000000')

        final_text = ['Внимание!', 'Оплата данного счета означает согласие с условиями поставки товара.',
                      'Уведомление об оплате обязательно, в противном случае не гарантируется наличие товара на складе.',
                      'Товар отпускается по факту прихода денег на р/с Поставщика, самовывозом,'
                      ' при наличии доверенности и паспорта.']

        row += 2

        for i in range(4):
            sheet[f'B{row + i}'].font = Font('Arial', 9)
            sheet.merge_cells(f'B{row + i}:AQ{row + i}')
            sheet.row_dimensions[row + i].height = k
            sheet[f'B{row + i}'] = final_text[i]
            sheet[f'B{row + i}'].alignment = Alignment(horizontal='left',
                                                       vertical='top',
                                                       text_rotation=0,
                                                       wrap_text=True, indent=0)

        sheet.row_dimensions[row + i].height = k * 2

        row += i + 1

        for letter in arr:
            sheet[f'{letter}{row}'].border = Border(bottom=Side(border_style=thick, color='FF000000'))

        row += 2
        sheet[f'B{row}'] = 'Руководитель'
        sheet[f'B{row}'].alignment = Alignment(horizontal='left',
                                               vertical='top',
                                               text_rotation=0,
                                               wrap_text=False,
                                               shrink_to_fit=False, indent=0)
        sheet[f'B{row}'].font = Font('Arial', 9, bold=True)

        sheet[f'Z{row}'] = 'Бухгалтер'
        sheet[f'Z{row}'].alignment = Alignment(horizontal='left',
                                               vertical='top',
                                               text_rotation=0,
                                               wrap_text=False,
                                               shrink_to_fit=False, indent=0)
        sheet[f'Z{row}'].font = Font('Arial', 9, bold=True)
        for letter in ['H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:
            sheet[f'{letter}{row}'].border = Border(bottom=Side(border_style=thin, color='FF000000'))

        for letter in ['AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ']:
            sheet[f'{letter}{row}'].border = Border(bottom=Side(border_style=thin, color='FF000000'))

    def get_root(path, tree, n=-1):
        path = path.split(',')
        root = tree.getroot()
        flag = False
        arr = []
        for step in path[:-1]:
            if '(' in step:
                if '=' in step:
                    step, arg = step.split('(')
                    arg = arg[:-1]
                    param, val = arg.split('=')
                elif 'Plur' in step:
                    step, arg = step.split('(')
                    flag = True
            a = [x.tag for x in root]
            if arr:
                while arr:
                    if a.count(arr[0]):
                        root = root[a.index(arr[0])]
                        break
                    del arr[0]
            else:
                count = 0
                for el in a:
                    if el == step:
                        count += 1
                if flag:
                    indexes = []
                    for j in range(len(a)):
                        if a[j] == step:
                            indexes.append(j)
                    root = root[indexes[n]]
                    flag = False
                elif count == 1:
                    root = root[a.index(step)]
                elif count > 1:
                    indexes = []
                    for j in range(len(a)):
                        if a[j] == step:
                            indexes.append(j)
                    for j in indexes:
                        try:
                            if root[j].attrib[param] == val:
                                root = root[j]
                        except IndexError:
                            pass
                        except KeyError:
                            pass
        if '/' in path[-1]:
            arr = path[-1].split('/')
            while arr:
                try:
                    return root.attrib[arr[0].strip()]
                except KeyError:
                    del arr[0]
            return ''
        try:
            return root.attrib[path[-1].strip()]
        except KeyError:
            return ''

    k = 13
    thin = 'thin'
    thick = 'medium'
    shutil.copyfile('ЭДО\\Schet.xlsx', f'{result_file}.xlsx')
    workbook = load_workbook('ЭДО\\Schet.xlsx')
    sheet = workbook['TDSheet']

    i = 0

    cols = []

    string = False
    table = False

    tree = ET.parse(
        open_filename)

    f = open('config.txt', encoding='utf-8').readlines()
    while i < len(f):
        el = f[i].strip()
        if 'ТАБЛИЦА' in el:
            s, cell = el.split(';')
            while not cell.isdigit():
                cell = cell[1:]
            row = int(cell)
            i += 1
            el = f[i].strip()
            while 'ТАБЛИЦА' not in el:
                route, col = el.split(';')
                cols.append((route, col))
                i += 1
                el = f[i].strip()
            create_table_end(row, cols, tree)
            break
        elif 'СТРОКА' in el:
            route, cell = el.split(';')
            s = ''
            i += 1
            el = f[i].strip()
            temp = el.split(';')
            for part in temp:
                if '"' in part:
                    part = part.replace('"', '')
                    s += part
                else:
                    s += get_root(part, tree)
            sheet[cell] = s
            i += 1
        else:
            route, cell = el.split(';')
            sheet[cell] = get_root(route, tree)
        i += 1

    workbook.save(f'{result_file}.xlsx')


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        uic.loadUi('MainWindow.ui', self)
        self.fileButton_3.clicked.connect(self.choseFile)
        self.pushButton.clicked.connect(self.tryConvert)
        self.label_2.setText('')

    def choseFile(self):
        fname = QFileDialog.getOpenFileName(self, 'Выбрать файл', '', 'Файл (*.xml)')[0]
        self.lineEdit.setText(fname)
        self.label_2.setText('')

    def tryConvert(self):
        if self.lineEdit.text() != '':
            start_file = self.lineEdit.text()
            result_file = '.'.join(start_file.split('.')[:-1])
            convert(start_file, result_file)
            self.label_2.setText(f'Файл {result_file.split("/")[-1]}\nготов')
        elif self.lineEdit.text() == '':
            self.label_2.setText('Выберите исходный файл.')



app = QtWidgets.QApplication([])
application = MainWindow()
application.show()

sys.exit(app.exec())
